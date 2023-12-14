from django.shortcuts import render, reverse, redirect,get_object_or_404
#from voting.models import Voter, Position, Candidate, Votes
from Users.models import User
from Users.forms import*
#from voting.forms import *
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from Users.views import account_login
from django.template.loader import render_to_string
from django.contrib import messages
from django.conf import settings
import requests
import json
from Users.models import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from.forms import*
from loan.models import*
from django.db.models import Sum
from reportlab.pdfgen import canvas
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from django import template
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
register = template.Library()
#from paypalrestsdk import Payment
import paypalrestsdk
user = get_user_model()
from django.shortcuts import get_object_or_404

from datetime import datetime



#TO DISPLAY THE HOME PAGE AND THE TEAM OF ZETECH


def index(request):
    if not request.user.is_authenticated:
        return account_login(request)
    context = {}


@login_required(login_url='account_login')
def creategroup(request):
    user = request.user
    members = Members.objects.filter(user=user)
    
    groups = set()  # Use a set to avoid duplicate groups
    
    for member in members:
        groups.update(member.groups.all())
    
    form = GroupForm()

    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            new_group = form.save(commit=False)
            new_group.Chairperson = request.user
            new_group.save()

            # Create only one Members object for the user
            member=Members.objects.create(user=user)
            member.groups.add(new_group)

            messages.success(request, "New Group Created")
            return redirect('add_group')
        else:
            messages.error(request, "Form errors")
   
    context = {
        'groups': groups if groups else None,
        'form1': form,
    }
    return render(request, "Main/Create_group.html", context)




@login_required(login_url='account_login')
def dashboard(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group
    total_members = members.count()  # Count the total number of members

    # Get cash contributions made by all members of the group
    group_cash_contributions = Cash.objects.filter(groups=groups)
    # Get contributions made by all members of the group
    group_contributions = Contribution.objects.filter(groups=groups)

    # Calculate the total contribution
    total_contribution = group_contributions.aggregate(Sum('amount'))['amount__sum'] or 0

    # Calculate the total cash collected
    total_cash_collected = group_cash_contributions.aggregate(Sum('amount'))['amount__sum'] or 0
    # Calculate the combined total amount
    combined_total = total_contribution + total_cash_collected

    # Get all expenses related to the group
    group_expenses = expenses.objects.filter(groups=groups)

    # Calculate the total sum of expenses
    total_expenses = group_expenses.aggregate(Sum('amount'))['amount__sum'] or 0

    # Get total pending loans for the group
    total_pending_loans = Loan.objects.filter(groups=groups, loan_status=0).count()

    # Get total approved loans for the group
    total_approved_loans = Loan.objects.filter(groups=groups, loan_status=1).count()
    
    # Get total Paid loans for the group
    Total_paid_loans =Loan.objects.filter(groups=groups,loan_status=3).count()

    # Get all loans in the group
    group_loans = Loan.objects.filter(groups=groups)

    # Calculate the sum of all paid amounts and remaining amount for the group
    total_paid_amount = sum([loan.calculate_total_amount() for loan in group_loans if loan.get_payment_status() == 'Fully Paid'])
    remaining_amount_to_pay = sum([loan.calculate_remaining_amount() for loan in group_loans if loan.get_payment_status() != 'Fully Paid'])
 



    context = {
        'groups': groups,
        'total_members': total_members,
        'total_contribution': total_contribution,
        'total_cash_collected': total_cash_collected,
        'combined_total': combined_total,
        'total_expenses': total_expenses,
        'total_pending_loans': total_pending_loans,
        'total_approved_loans': total_approved_loans,
        'Total_paid_loans':Total_paid_loans,
        'total_paid_amount':total_paid_amount,
        'remaining_amount_to_pay':  remaining_amount_to_pay,
        'members' : members ,

    }

    return render(request, "loan/Home.html", context)





@login_required(login_url='account_login')
def add_selected_members(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')
        selected_users = User.objects.filter(id__in=selected_user_ids)

        for user in selected_users:
            # Check if the user is already a member of the selected group
            existing_membership = Members.objects.filter(user=user, groups=group).first()

            if existing_membership:
                messages.warning(request, f'{user.username} is already a member of this group.')
            else:
                # Create a new membership for the user and add it to the group
                member = Members(user=user)
                member.save()
                group.membership.add(member)
                messages.success(request, f'{user.username} has been added to the group successfully.')

        return redirect('view_members', group_id=group.id)



@login_required(login_url='account_login')
def view_members(request, group_id):
    groups = get_object_or_404(Group, id=group_id)
    members = Members.objects.filter(groups=groups)
    all_users = User.objects.all()

    if request.method == 'POST':
        if 'remove_member' in request.POST:
            # Handle the removal of a member
            member_id = request.POST.get('remove_member')
            member = get_object_or_404(Members, id=member_id)
            member.delete()
            messages.success(request, 'Member removed successfully.')
            return redirect('view_members', group_id=groups.id)

        member_form = MemberForm(request.POST)
        user_form = userForm(request.POST)

        if member_form.is_valid() and user_form.is_valid():
            user = user_form.save()

            # Check if the user is already a member of the group
            existing_member = Members.objects.filter(user=user, groups=groups).first()

            if existing_member:
                messages.error(request, f'{user.username} is already a member of this group.')
            else:
                member = member_form.save(commit=False)
                member.user = user
                member.save()

                groups.membership.add(member)
                messages.success(request, f'{user.username} has been added to the group successfully.')

            return redirect('view_members', group_id=groups.id)
    else:
        member_form = MemberForm()
        user_form = userForm()

    context = {
        'groups': groups,
        'member_form': member_form,
        'user_form': user_form,
        'members': members,
        'all_users': all_users,
    }

    return render(request, "loan/members.html", context)


@login_required(login_url='account_login')
def iew_members(request, group_id):
    groups = get_object_or_404(Group, id=group_id)
    members = Members.objects.filter(groups=groups)
    all_users = User.objects.all()


    if request.method == 'POST':
        member_form = MemberForm(request.POST)
        user_form = userForm(request.POST)

        if member_form.is_valid() and user_form.is_valid():
            user = user_form.save()

            # Check if the user is already a member of the group
            existing_member = Members.objects.filter(user=user, groups=groups).first()

            if existing_member:
                messages.error(request, f'{user.username} is already a member of this group.')
            else:
                member = member_form.save(commit=False)
                member.user = user
                member.save()

                groups.membership.add(member)
                messages.success(request, f'{user.username} has been added to the group successfully.')

            return redirect('view_members', group_id=groups.id)
    else:
        member_form = MemberForm()
        user_form = userForm()

    context = {
        'groups': groups,
        'member_form': member_form,
        'user_form': user_form,
        'members': members,
        'all_users': all_users,

    }
    return render(request, "loan/members.html", context)


#original view member
@login_required(login_url='account_login')
def vie_members(request, group_id):
    groups = get_object_or_404(Group, id=group_id)
    members = Members.objects.filter(groups=groups)

    if request.method == 'POST':
        member_form = MemberForm(request.POST)
        user_form = userForm(request.POST)

        if member_form.is_valid() and user_form.is_valid():
            # Save the user information
            user = user_form.save()

            # Save the member information
            member = member_form.save(commit=False)
            member.user = user
            member.save()

            # Add the member to the group
            groups.members.add(member)

            return redirect('view_members', group_id=groups.id)
    else:
        member_form = MemberForm()
        user_form = userForm()

    context = {
        'groups': groups,
        'member_form': member_form,
        'user_form': user_form,
        'members': members,
        'all_users': all_users,
    }
    return render(request, "loan/members.html", context)


@login_required(login_url='account_login')
def group_contributions(request, pk):
    groups = get_object_or_404(Group, id=pk)

    members = Members.objects.filter(user=request.user, groups=groups)
    member = members.first()
    contribution = Contribution.objects.filter(member__in=members)
    total_contribute = contribution.aggregate(Sum('amount'))['amount__sum'] or 0

    member_contributions = []
    for member in members:
        contribute = Contribution.objects.filter(member=member, groups=groups)
        total_contribution = Contribution.objects.filter(member=member).aggregate(Sum('amount'))['amount__sum'] or 0
        member_contributions.append({'member': member, 'total_contribution': total_contribution})

    contribute_form = ContributionForm()     
    if request.method == 'POST':
        if 'withdraw' in request.POST:
            withdrawal_amount = float(request.POST.get('withdrawal_amount', 0))
            withdrawal_pin = int(request.POST.get('withdrawal_pin',0))

            # Validate withdrawal PIN
            try:
                # Assuming there's only one member for simplicity
                member = Members.objects.get(user=request.user, groups=groups)
            except Members.DoesNotExist:
                messages.error(request, "Member not found.")
                return redirect('group_contributions', pk=pk)

            if withdrawal_pin != member.pin:
                messages.error(request, "Incorrect PIN. Please try again.")
                return redirect('group_contributions', pk=pk)

            # Validate withdrawal amount against group limits
            if (
                groups.withdrawal_minimum is not None
                and withdrawal_amount < groups.withdrawal_minimum
            ) or (
                groups.withdrawal_maximum is not None
                and withdrawal_amount > groups.withdrawal_maximum
            ):
                messages.error(request, "Withdrawal amount is outside group limits.")
                return redirect('group_contributions', pk=pk)

            # Validate withdrawal amount against member's total contribution
            if withdrawal_amount > total_contribute:
                messages.error(request, "Withdrawal amount exceeds total contribution.")
                return redirect('group_contributions', pk=pk)

            # Create a withdrawal contribution
            withdrawal_contribution = Contribution.objects.create(
                member=member, groups=groups, amount=-withdrawal_amount
            )

            messages.success(
                request,
                f"Withdrawal of {withdrawal_amount} successful. Updated total contribution: {total_contribute - withdrawal_amount}",
            )

            return redirect('group_contributions', pk=pk)
        else:
            contribute_form = ContributionForm(request.POST) 
            if contribute_form.is_valid():
                contribution = contribute_form.save(commit=False)
                contribution.member = member
                contribution.groups = groups
                contribute_form.save()
                contribute_form.save_m2m()
                messages.success(request, "The Payment is successfully and contribution added successfully")
                return redirect('group_contributions', pk=pk)
            else:
                messages.error(request, "Transaction is canceled")

            contribute_form = ContributionForm(request.POST)

    context = {
        'groups': groups,
        'member_contributions': member_contributions,
        'contribute_form': contribute_form,
        'members': members,
        'contribution': contribution,
        'total_contribute': total_contribute,
    }

    return render(request, "loan/Contribution.html", context)





@login_required(login_url='account_login')
def cash_collected(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)
    mb = Members.objects.filter(user=request.user, groups=groups)

    # Use the first member or handle multiple members appropriately
    member = mb.first()

    # Get cash collected by all members of the group
    group_cash = Cash.objects.filter(member__in=members,groups=groups)
    if request.method == 'POST':
        form = CashForm(request.POST)
        if form.is_valid():
            contribution = form.save(commit=False)
            contribution.member = member
            contribution.groups = groups
            contribution.save()
            form.save_m2m()
            messages.success(request, "Cash added successfully")
            return redirect( 'cash_collected',pk=pk)
        else:
            messages.error(request, "Form errors")
           
    else:
        form =CashForm()

    # Create a list to store member-wise totals
    member_totals = []

    # Calculate the total cash collected for each member and add to the list
    for member in members:
        member_cash = group_cash.filter(member=member)
        member_total_cash = member_cash.aggregate(Sum('amount'))['amount__sum'] or 0
        member_totals.append({'member': member,'total_cash': member_total_cash,})
  

    # Calculate the total cash collected for the entire group
    total_cash = group_cash.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'groups': groups,
        'members': member_totals,  # Pass the member-wise totals to the template
        'total_cash': total_cash,
        'contribution_form': form,
        'members' : members ,
        'member_totals':member_totals,

    }

    return render(request, "loan/cash_collected.html", context)





@login_required(login_url='account_login')
def expense_view(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group
    total_members = members.count()  # Count the total number of members

    # Get all expenses related to the group
    group_expenses = expenses.objects.filter(groups=groups)

    if request.method == 'POST':
        form = ExpensesForm(request.POST)
        if form.is_valid():
            contribution = form.save(commit=False)
            #contribution.groups.set([groups])
            #contribution.groups= Group.objects.get(groups=groups)
            contribution.save()
             # Make sure to save the instance before adding the group
            form.save_m2m()  

            contribution.groups.set([groups]) 
            messages.success(request, "Group expense added successfully")
            return redirect( 'group_expenses',pk=pk)
        else:
            messages.error(request, "Form errors")
           
    else:
        form =ExpensesForm()

    # Calculate the total sum of expenses
    total_expenses = group_expenses.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'groups': groups,
        'total_expenses': total_expenses,
        'group_expenses':group_expenses,
        'contribution_form': form,
        'members' : members ,
       

    }
    return render(request, "loan/expenses.html", context)


@login_required(login_url='account_login')
def view_approved_loan(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group


    # Get total approved loans for the group
    total_approved_loans = Loan.objects.filter(groups=groups, loan_status=1).count()
    approved_loans = Loan.objects.filter(groups=groups, loan_status=1)

    context = {
        'groups': groups,
        'total_approved_loans': total_approved_loans,
        'approved_loans': approved_loans,
        'members' : members ,

    }

    return render(request, "loan/loan_approved.html", context)


@login_required(login_url='account_login')
def group_contributions_list(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)
    
    member_contributions = []
    total_group_contribution = 0
    
    for member in members:
        contributions = Contribution.objects.filter(member=member, groups=groups)
        total_contribution = contributions.aggregate(Sum('amount'))['amount__sum'] or 0
        total_withdrawal = contributions.filter(amount__lt=0).aggregate(Sum('amount'))['amount__sum'] or 0
        total_group_contribution += total_contribution
        
        member_contributions.append({
            'member': member,
            'total_contribution': total_contribution,
            'contributions':contributions,
            'total_withdrawal': abs(total_withdrawal),
        })
    # Generate Excel file
    #excel_file_path = generate_excel(groups, member_contributions, total_group_contribution)


    context = {
        'groups': groups,
        'member_contributions': member_contributions,
        'total_group_contribution': total_group_contribution,
    }

    return render(request, "Main/total_contribution_group.html", context)

def download_excel(request, pk):
    groups = get_object_or_404(Group, id=pk)
    member_contributions = []
    total_group_contribution = 0
    
    # Populate member_contributions and total_group_contribution with actual data
    for member in Members.objects.filter(groups=groups):
        contributions = Contribution.objects.filter(member=member, groups=groups)
        total_contribution = contributions.aggregate(Sum('amount'))['amount__sum'] or 0
        total_withdrawal = contributions.filter(amount__lt=0).aggregate(Sum('amount'))['amount__sum'] or 0
        total_group_contribution += total_contribution
        
        member_contributions.append({
            'member': member,
            'total_contribution': total_contribution,
            'contributions': contributions,
            'total_withdrawal': abs(total_withdrawal),
        })

    # Generate Excel file
    excel_file_path = generate_excel(groups, member_contributions, total_group_contribution)

    with open(excel_file_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{excel_file_path}"'

    return response

def generate_excel(groups, member_contributions, total_group_contribution):
    # Check if groups is None
    if groups is None:
        raise ValueError("Invalid groups object")

    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Set the group name at the top of the sheet
    sheet['A1'] = f"Group: {groups.Name}"
    sheet['A1'].font = Font(size=14, bold=True)

    # Write column headers
    headers = ["Date_contributed", "Full_Name", "Category", "amount"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=3, column=col_num, value=header).font = Font(bold=True)

    # Write member contributions
    row_num = 4
    for member_contribution in member_contributions:
        for contribution in member_contribution['contributions']:
            # Ensure contribution.date_created is a valid datetime object
            if contribution.date_created and isinstance(contribution.date_created, datetime):
                # Convert to a string with a format that Excel can handle
                date_created_str = contribution.date_created.strftime('%Y-%m-%d %H:%M:%S')

                # Write the formatted string to the Excel cell
                sheet.cell(row=row_num, column=1, value=date_created_str)
            else:
                # Handle the case where date_created is None or not a datetime object
                sheet.cell(row=row_num, column=1, value="Invalid Date")

            sheet.cell(row=row_num, column=2, value=f"{member_contribution['member'].user.first_name} {member_contribution['member'].user.last_name}")
            
            # Check if contribution.categories is not None before accessing its 'Name' attribute
            category_name = contribution.categories.Name if contribution.categories else 'Unknown'
            sheet.cell(row=row_num, column=3, value=category_name)
            
            sheet.cell(row=row_num, column=4, value=f"{contribution.amount} KES")
            row_num += 1

        # Add a row for the total amount of the member
        sheet.cell(row=row_num, column=1, value="Total member amount:")
        sheet.cell(row=row_num, column=4, value=f"{member_contribution['total_contribution']} KES")
        row_num += 1

    # Add a row for the total group contribution
    sheet.cell(row=row_num, column=1, value="Total Group Contribution:")
    sheet.cell(row=row_num, column=4, value=f"{total_group_contribution} KES")

    # Save the workbook
    file_path = f"{groups.Name}_contributions.xlsx"
    workbook.save(file_path)

    return file_path



@login_required(login_url='account_login')
def view_pending_loan(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group

    pending_loans = Loan.objects.filter(groups=groups, loan_status=0)
       # Get total pending loans for the group
    total_pending_loans = Loan.objects.filter(groups=groups, loan_status=0).count()

    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')
        new_status = request.POST.get('new_status')
        
        loan = get_object_or_404(Loan, id=loan_id, groups=groups, loan_status=0)

        if request.user == groups.Chairperson:
            # Update the loan status based on the button clicked
            if new_status == 'approved':
                loan.loan_status = 1  # 1 represents approved status, adjust based on your model
                messages.success(request, 'The Loan is  has been approved successfully.')
            elif new_status == 'rejected':
                loan.loan_status = 2  # 2 represents rejected status, adjust based on your model
                messages.success(request, 'The Loan is  has been disapproved successfully.')
            loan.save()


    context = {
        'groups': groups,
        'total_pending_loans': total_pending_loans,
        'pending_loans': pending_loans,
        'members' : members ,

    }

    return render(request, "loan/loan_pending.html", context)



@login_required(login_url='account_login')
def create_view_fines(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group
    fines = fine.objects.filter(groups=groups)

    if request.method == 'POST':
        form = FineForm(request.POST)
        if form.is_valid():
            contribution = form.save(commit=False)
            #contribution.groups.set([groups])
            #contribution.groups= Group.objects.get(groups=groups)
            contribution.save()
             # Make sure to save the instance before adding the group
            form.save_m2m()  

            contribution.groups.set([groups]) 
            messages.success(request, "Group Fine added successfully")
            return redirect( 'group_fines',pk=pk)
        else:
            messages.error(request, "Form errors")
           
    else:
        form =FineForm()

    total_fines = fines.aggregate(Sum('amount'))['amount__sum'] or 0
    context = {
        'groups': groups,
        'fines':fines,
        'contribution_form': form,
        'total_fines':total_fines,
        'members' : members ,
       

    }
    return render(request, "loan/fine.html", context)




@login_required(login_url='account_login')
def view_Paid_loan(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group

    # Get all fully paid loans for the group
    Paid_loans = Loan.objects.filter(groups=groups, loan_status=3)

    total_amount_group = sum([loan.calculate_total_amount() for loan in Paid_loans])
    total_loan_amounts = {loan.id: loan.calculate_total_amount() for loan in  Paid_loans }

    # Create a dictionary to store loan ids and their corresponding total amounts
    loan_details = []

    for loan in Paid_loans:
        total_amount_to_return=total_loan_amounts.get(loan.id, 0)
        loan_details.append({
            'loan':loan,
            'total_amount_to_return':total_amount_to_return,
            'status_display': loan.get_loan_status_display(),

            })


    context = {
        'groups': groups,
        'Paid_loans': Paid_loans,
        'members': members,
        'loan_details':loan_details,
    }

    return render(request, "loan/paid_loan.html", context)





@login_required(login_url='account_login')
def view_Group_info(request, pk):
    groups= get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)

    if request.method == 'POST':
        form = GroupEditForm(request.POST,request.FILES, instance=groups)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your Group Profile has been updated successfully.')
            return redirect('view_Group', pk=pk)
        else:
            messages.error(request, 'There was an error updating your Group Profile. Please correct the errors below.')
    else:
        form = GroupEditForm(instance=groups)

    context = {
        'groups': groups,
        'members': members,
        'form': form,
    }

    return render(request, 'loan/group_information.html', context)






@login_required(login_url='account_login')
def view_Loan_Applications(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)

    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')
        new_status = request.POST.get('new_status')
        
        loan = get_object_or_404(Loan, id=loan_id, groups=groups, loan_status=0)

        if request.user == groups.Chairperson:
            # Update the loan status based on the button clicked
            if new_status == 'approved':
                loan.loan_status = 1  # 1 represents approved status, adjust based on your model
                messages.success(request, 'The Loan is  has been approved successfully.')
            elif new_status == 'rejected':
                loan.loan_status = 2  # 2 represents rejected status, adjust based on your model
                messages.success(request, 'The Loan is  has been disapproved successfully.')
            loan.save()

    pending_loans = Loan.objects.filter(groups=groups, loan_status=0)
    total_pending_loans = Loan.objects.filter(groups=groups, loan_status=0).count()

    context = {
        'groups': groups,
        'total_pending_loans': total_pending_loans,
        'pending_loans': pending_loans,
        'members': members,
    }
    if 'download_excel' in request.GET:
        # Generate Excel
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{groups.Name}_loan_applications.xlsx"'

        # Create Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Add group name to Excel
        worksheet.merge_cells('A1:E1')
        cell = worksheet['A1']
        cell.value = f'Group: {groups.Name}'
        cell.font = Font(size=14, bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # Create header row
        header_row = ["Date Applied", "Full Name", "Period of Payment", "Loan Status", "Borrowed Amount (Ksh)"]
        worksheet.append(header_row)

        # Add loan information to Excel
        for loan in pending_loans:
            row = [str(loan.date_applied), f"{loan.member.user.first_name} {loan.member.user.last_name}",
                   str(loan.duration_months), loan.get_loan_status_display(), str(loan.amount)]
            worksheet.append(row)

        # Save the workbook to the response
        workbook.save(response)

        return response

    return render(request, "loan/group_loan_applications.html", context)

    
@login_required(login_url='account_login')
def group_accounts(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)  # Filter members by the specific group

    # Get cash contributions made by all members of the group
    group_cash_contributions = Cash.objects.filter(groups=groups)
    # Get contributions made by all members of the group
    group_contributions = Contribution.objects.filter(groups=groups)

    # Calculate the total contribution
    total_contribution = group_contributions.aggregate(Sum('amount'))['amount__sum'] or 0

    # Calculate the total cash collected
    total_cash_collected = group_cash_contributions.aggregate(Sum('amount'))['amount__sum'] or 0
    # Calculate the combined total amount
    combined_total = total_contribution + total_cash_collected

    
    # Get total Paid loans for the group
    Total_paid_loans =Loan.objects.filter(groups=groups,loan_status=3).count()

    # Get all loans in the group
    group_loans = Loan.objects.filter(groups=groups)

    # Calculate the sum of all paid amounts and remaining amount for the group
    total_paid_amount = sum([loan.calculate_total_amount() for loan in group_loans if loan.get_payment_status() == 'Fully Paid'])
    remaining_amount_to_pay = sum([loan.calculate_remaining_amount() for loan in group_loans if loan.get_payment_status() != 'Fully Paid'])
    # Calculate the total interest for fully paid loans in the group
    total_int = sum([loan.calculate_interest() for loan in group_loans if loan.get_payment_status() == 'Fully Paid'])
    # Calculate the total interest for the group
    total_interest = sum([loan.calculate_interest() for loan in group_loans])



    context = {
        'groups': groups,
        'total_contribution': total_contribution,
        'total_cash_collected': total_cash_collected,
        'combined_total': combined_total,
        'Total_paid_loans':Total_paid_loans,
        'total_paid_amount':total_paid_amount,
        'remaining_amount_to_pay':  remaining_amount_to_pay,
        'members' : members ,
        'total_interest':total_interest,
        'total_int':total_int,

    }

    return render(request, "Main/group_balances.html", context)





    


@login_required(login_url='account_login')
def create_pin(request, pk):
    groups = get_object_or_404(Group, id=pk)
    if request.method == 'POST':
        form = PinCreationForm(request.POST)
        if form.is_valid():
            pin =form.cleaned_data['pin'] 
            
            # Assuming you have a logged-in user
            user = request.user
            group_membership = groups.membership.get(user=user)
            group_membership.pin = pin
            group_membership.save()
            messages.success(request, 'The Pin has been set successfully.')
            # Redirect to a success page or perform additional actions
            return redirect('user_profile',pk=pk)
        else:
            messages.error(request,'The pin is invalid')

    else:
        form = PinCreationForm()

    return render(request, 'Main/create_pin.html', {'form': form, 'groups': groups})