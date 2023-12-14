from django.shortcuts import render, reverse, redirect,get_object_or_404
#from voting.models import Voter, Position, Candidate, Votes
from Users.models import User
from Users.forms import*
#from voting.forms import *
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from Users.views import account_login
from django.template.loader import render_to_string
from django.contrib import messages
from django.conf import settings
import requests
import json
from Users.models import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from loan.models import*
from django.db.models import Sum
from reportlab.pdfgen import canvas
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from Main.forms import*
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from django import template
register = template.Library()
from django.utils.timezone import now
#from paypalrestsdk import Payment
import paypalrestsdk
user = get_user_model()
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from io import BytesIO
from reportlab.pdfgen import canvas
from datetime import datetime, timedelta, timezone
from django.utils.dateformat import format as date_format
from openpyxl import Workbook
from django.db import transaction
import pandas as pd
from django.http import HttpResponse
from io import BytesIO

@login_required(login_url='account_login')
def view_Paid_loan_excel(request, pk):
    groups = get_object_or_404(Group, id=pk)
    Paid_loans = Loan.objects.filter(groups=groups, loan_status=3)

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{groups.Name}_paid_loans.xlsx"'

    # Create Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Add group name to Excel
    worksheet.merge_cells('A1:F1')
    cell = worksheet['A1']
    cell.value = f'Group: {groups.Name}'
    cell.font = Font(size=18, bold=True)
    cell.alignment = Alignment(horizontal='center')

    # Add headers
    headers = ["Date_Paid", "Full Name", "Period", "Interest Rate/PM", "Borrowed Amount", "Status", "Total Amount", "Interest"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}2'] = header
        worksheet[f'{col_letter}2'].font = Font(size=14, bold=True)
        worksheet.column_dimensions[col_letter].width = 15  # Adjust column width
      
    # Add loan details
    total_interest_paid = 0  # Initialize total interest for paid loans
    for row_num, contribution in enumerate(Paid_loans, 3):
        # Calculate interest for each contribution
        interest_for_contribution = contribution.calculate_interest()
        worksheet.cell(row=row_num, column=8, value=f"{round(interest_for_contribution, 2)} Ksh")

        # Accumulate the total interest for paid loans
        total_interest_paid += interest_for_contribution

        worksheet.cell(row=row_num, column=1, value=contribution.payment_set.first().payment_date.strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.cell(row=row_num, column=2, value=f"{contribution.member.user.first_name} {contribution.member.user.last_name}")
        worksheet.cell(row=row_num, column=3, value=contribution.duration_months)
        worksheet.cell(row=row_num, column=4, value=f"{groups.loan_interest_rate}%")
        worksheet.cell(row=row_num, column=5, value=f"{contribution.amount} Ksh")
        worksheet.cell(row=row_num, column=6, value=contribution.get_loan_status_display())
        worksheet.cell(row=row_num, column=7, value=f"{round(contribution.calculate_total_amount(), 2)} Ksh")  # Round to two decimal places

    # Add row for total sum in the center
    total_sum_row = row_num + 1
    total_sum = round(sum(contribution.calculate_total_amount() for contribution in Paid_loans), 2)  # Round the total sum

    worksheet.cell(row=total_sum_row, column=1, value=f'Total Sum:')
    worksheet.merge_cells(start_row=total_sum_row, start_column=1, end_row=total_sum_row, end_column=6)
    worksheet['A' + str(total_sum_row)].alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=total_sum_row, column=7, value=f'{total_sum} Ksh')  # Use the rounded total sum
    worksheet.cell(row=total_sum_row, column=8, value=f'{total_interest_paid} Ksh')  # Use the rounded total interest for paid loans
    worksheet['G' + str(total_sum_row)].alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(response)
    return response






def get_member_loans_and_total_amount(member_id, group_id):
    # Get the group and member
    groups = get_object_or_404(Group, id=group_id)
    member = get_object_or_404(Members, id=member_id, groups=groups)


    # Get all loans in the group for the specific member
    #loans = Loan.objects.filter(groups=groups, member=member)
    member_loans = Loan.objects.filter(member__id=member_id, groups__id=group_id)

    # Calculate the total amount borrowed by the member
    #total_loans = member_loans.count()
    total_amount_to_return = sum([loan.calculate_total_amount() for loan in member_loans])
    #total_amount_to_return = member_loans.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    #total_borrowed_amount = sum([loan.calculate_total_amount() for loan in loans if loan.calculate_total_amount() is not None])

    # Return the loans and total amount
    return  member_loans , total_amount_to_return

   


@login_required(login_url='account_login')
@transaction.atomic
def apply_loan(request, pk):
    groups = get_object_or_404(Group, id=pk)
    #member = Members.objects.get(user=request.user, groups=groups)
    member = Members.objects.filter(user=request.user, groups=groups)

    # Use the first member or handle multiple members appropriately
    member = member.first()

    # Assuming you want to handle the case where multiple members are found
    # Get specific member's loans and total amount
    member_loans, total_member_amount = get_member_loans_and_total_amount(member_id=member.id, group_id=pk)

    # Calculate each member's total loans and total amount to return in the group
    #total_loans, total_amount_to_return = get_member_total_loans_and_amount(member_id=member.id, group_id=pk)
    member_loans , total_amount_to_return = get_member_loans_and_total_amount(member_id=member.id, group_id=pk)
    
    # Calculate the total amount of each loan applied by the member
    total_loan_amounts = {loan.id: loan.calculate_total_amount() for loan in member_loans}

    # Include total amount, remaining amount, and other loan details in the context
    loan_details = []
    for loan in member_loans:
        total_amount_to_return = total_loan_amounts.get(loan.id, 0)
        remaining_amount = max(0, loan.calculate_remaining_amount())
        loan_details.append({
            'loan': loan,
            'total_amount_to_return': total_amount_to_return,
            'remaining_amount': remaining_amount,
            'status_display': loan.get_loan_status_display()
        })
   

    # Include total amount for each loan in the context
    #loan_details = [{'loan': loan, 'total_amount_to_return': total_loan_amounts.get(loan.id, 0)} for loan in member_loans]

    if request.method == 'POST':
        form = LoanApplicationForm(request.POST)
        payment_form = PaymentForm(request.POST)

        if form.is_valid():
            loan = form.save(commit=False)
            loan.member = member
            #loan.member = Members.objects.get(user=request.user)
            loan.groups = groups
            total_amount = loan.calculate_total_amount()
            loan.total_amount = total_amount
             #if not member_loans or any
             #if not member_loans and any
            # Check if the member has no existing loans and if there are any loans with status 3 or 4
            if any(loan.loan_status in [0,2 ] for loan in member_loans):
                messages.warning(request,  'You cannot apply for a new loan while you have a pending or approved loan.')
                return redirect('apply_loans', pk=pk)
            elif any(loan.loan_status in [0, 1] for loan in member_loans):
                messages.warning(request,'You cannot apply for a new loan until your existing loan is fully paid or partially paid.')
                return redirect('apply_loans', pk=pk)
            else:
                loan.save()
                form.save_m2m()
                messages.success(request, f'Loan application submitted successfully! Total loan repayment amount: {total_amount} Ksh')
                return redirect('apply_loans', pk=pk)
            
            # If the loan is approved, update the member's account balance
            if loan.loan_status == 1:
                member.account_balance += loan.total_amount
                member.save()
                messages.success(request, f'Loan approved! Amount of {loan.total_amount} Ksh added to your account balance.')

            messages.success(request, f'Loan application submitted successfully! Total loan repayment amount: {total_amount} Ksh')
            return redirect('apply_loans', pk=pk)

        elif payment_form.is_valid():
            amount_paid = payment_form.cleaned_data['amount_paid']
            loan_id = request.POST.get('loan_id')
            loan = get_object_or_404(Loan, id=loan_id)

          # Check if the loan status is approved and partially paid before allowing payment
            if loan.loan_status != 1 and loan.loan_status != 4:
                messages.error(request, 'You can only pay approved and partially paid loans!')
                return redirect('apply_loans', pk=pk)



            # Check if the payment exceeds the total loan amount
            if amount_paid > loan.calculate_total_amount():
                messages.error(request, 'You are paying more than your total loan amount!')
                return redirect('apply_loans', pk=pk)

            
            # Make datetime.now() timezone-aware
            current_datetime = datetime.now(timezone.utc)

            # Ensure loan.date_applied is timezone-aware
            if loan.date_applied.tzinfo is None or loan.date_applied.tzinfo.utcoffset(loan.date_applied) is None:
                loan.date_applied = loan.date_applied.replace(tzinfo=timezone.utc)

            # Check if the payment is made within the loan's duration
            if current_datetime > loan.date_applied + timedelta(days=loan.duration_months * 30):
                messages.error(request, 'The loan payment is past the due date. Late payments are not allowed.')
                return redirect('apply_loans', pk=pk)




            # Update the remaining amount based on the payment
            remaining_amount = max(0, loan.calculate_remaining_amount())

            

            # Check if the loan is fully paid or partially paid
            if remaining_amount <= 0:
                loan.loan_status = 4  # Fully paid
                loan.is_fully_paid = True
            else:
                loan.loan_status = 3  # Partially paid
                loan.is_fully_paid = False


            # Check if the loan is fully paid
            if sum(payment.amount_paid for payment in Payment.objects.filter(loan=loan)) + amount_paid >= loan.calculate_total_amount():

                loan.loan_status = 3 
                loan.is_fully_paid = True

            else:
                loan.is_fully_paid = False
                loan.loan_status = 4
           # Save the loan status
          
            loan.save()

            # Record the payment
            Payment.objects.create(loan=loan, amount_paid=amount_paid)
            

            #loan.pay_loan(amount_paid)
            messages.success(request, f'Loan payment of {amount_paid} Ksh successfully processed!')

            # If the loan is fully paid, prevent further payments
            if loan.is_fully_paid:
                messages.warning(request, 'This loan is fully paid. No further payments allowed.')

            return redirect('apply_loans', pk=pk)

    else:
        form = LoanApplicationForm(initial={'groups': groups})
        payment_form = PaymentForm()

    context = {
        'groups': groups,
        'member': member,
        'form': form,
        'member_loans': member_loans,
        'total_amount_to_return': total_amount_to_return,
        'payment_form': payment_form,
        ' total_loan_amounts ': total_loan_amounts, 
        'loan_details': loan_details, 
    }

    return render(request, "loan/apply_loan.html", context)




@login_required(login_url='account_login')
def view_rejected_loans(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)

    rejected_loans = Loan.objects.filter(groups=groups, loan_status=2)  # Filter rejected loans
    total_rejected_loans = rejected_loans.count()

    if request.method == 'POST':
        if 'delete_all' in request.POST:
            # Delete all rejected loans
            rejected_loans.delete()
            messages.success(request, 'All rejected loans have been deleted successfully.')
            
        elif 'approve_all' in request.POST:
            # Approve all rejected loans
            for loan in rejected_loans:
                loan.loan_status = 1  # 1 represents approved status, adjust based on your model
                loan.save()
            messages.success(request, 'All rejected loans have been approved successfully.')
        
        elif 'approve_specific' in request.POST:
            # Get the loan ID from the form data
            loan_id = request.POST.get('loan_id')
            # Approve the specific rejected loan
            specific_loan = get_object_or_404(Loan, id=loan_id, groups=groups, loan_status=2)
            specific_loan.loan_status = 1  # 1 represents approved status, adjust based on your model
            specific_loan.save()
            messages.success(request, 'The selected loan has been approved successfully.')

    context = {
        'groups': groups,
        'total_rejected_loans': total_rejected_loans,
        'rejected_loans': rejected_loans,
        'members': members,
    }

    return render(request, "loan/rejected_loans.html", context)







@login_required(login_url='account_login')
def generate_excel_loan_details(request, pk):
    groups = get_object_or_404(Group, id=pk)
    member = Members.objects.get(user=request.user, groups=groups)
     #Get specific member's loans and total amount
    member_loans, total_member_amount = get_member_loans_and_total_amount(member_id=member.id, group_id=pk)
    # Calculate the total amount of each loan applied by the member
    total_loan_amounts = {loan.id: loan.calculate_total_amount() for loan in member_loans}

    # Include total amount, remaining amount, and other loan details in the context
    loan_details = []
    for loan in member_loans:
        total_amount_to_return = total_loan_amounts.get(loan.id, 0)
        remaining_amount = max(0, loan.calculate_remaining_amount())
        loan_details.append({
            'loan': loan,
            'total_amount_to_return': total_amount_to_return,
            'remaining_amount': remaining_amount,
            'status_display': loan.get_loan_status_display()
        })
   
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add header row
    header = [
        "Date of Application",
        "Full Name",
        "Period (Months)",
        "Interest Rate/PM (%)",
        "Borrowed Amount (Ksh)",
        "Status",
        "Total to Return Amount (Ksh)",
        "Remaining Amount (Ksh)",
    ]
    for col_num, header_text in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header_text
    # Add data rows
    for row_num, contribution in enumerate(loan_details, 2):
        data = [
            date_format(contribution['loan'].date_applied, 'Y-m-d H:i:s'),
            f"{contribution['loan'].member.user.first_name} {contribution['loan'].member.user.last_name}",
            contribution['loan'].duration_months,
            f"{contribution['loan'].groups.loan_interest_rate}%",
            f"{contribution['loan'].amount}",
            "Paid" if contribution['loan'].is_fully_paid else contribution['loan'].get_loan_status_display(),
            f"{contribution['total_amount_to_return']}",
            f"{contribution['remaining_amount']}",
        ]
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value


    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=loan_details_{pk}_{date_format(now(), "YmdHis")}.xlsx'
    workbook.save(response)

    return response







@login_required(login_url='account_login')
def investment_view(request, pk):
    groups = get_object_or_404(Group, id=pk)
    members = Members.objects.filter(groups=groups)

    # Get all expenses related to the group
    group_investment = Investment.objects.filter(groups=groups)

    if request.method == 'POST':
        form = InvestmentForm(request.POST)
        if form.is_valid():
            contribution = form.save(commit=False)
            contribution.save()
            form.save_m2m()
            contribution.groups.set([groups]) 
            messages.success(request, "Group Investment added successfully")
            return redirect('group_investment', pk=pk)
        else:
            messages.error(request, "Form errors")
    else:
        form = InvestmentForm()

    # Calculate the total sum of expenses
    total_investment = group_investment.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'groups': groups,
        'group_investment': group_investment,
        'contribution_form': form,
        'members': members,
    }

    if request.method == 'GET' and 'excel' in request.GET:
        # If the 'excel' parameter is present in the URL, generate Excel
        excel_buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add the group name at the top
        ws.append(["Group Name:", groups.Name])

        # Add headers to the Excel sheet
        headers = ["Name", "Investment Type", "Amount (KES)"]
        ws.append(headers)

        # Your loop for contribution data here...
        for contribution in group_investment:
            row_data = [contribution.Name, contribution.Investment_type, f"{contribution.amount} KES"]
            ws.append(row_data)

        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = HttpResponse(excel_buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="group_investment.xlsx"'
        return response

    return render(request, "loan/Group_investment.html", context)

@login_required(login_url='account_login')
def generate_excel_contribution_details(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    # Get all contributions made by members in the group
    contributions = Contribution.objects.filter(member__groups=group)

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add group name to Excel
    worksheet.merge_cells('A1:E1')
    cell = worksheet['A1']
    cell.value = f'Group: {group.Name}'
    cell.font = openpyxl.styles.Font(size=18, bold=True)
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # Add header row
    header = [
        "Date of Contribution",
        "Full Name",
        "Category",
        "Amount (Ksh)",
        # Add more fields as needed
    ]
    for col_num, header_text in enumerate(header, 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.value = header_text

    # Add data rows
    total_amount = 0
    for row_num, contribution in enumerate(contributions, 3):
        data = [
            date_format(contribution.date_created, 'Y-m-d H:i:s'),
            f"{contribution.member.user.first_name} {contribution.member.user.last_name}",
            str(contribution. categories),  # Convert Category to string
            f"{contribution.amount}",
            # Add more fields as needed
        ]
        total_amount += contribution.amount
        for col_num, cell_value in enumerate(data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Add total contribution row
    total_row_num = len(contributions) + 3
    worksheet.cell(row=total_row_num, column=1, value="Total Contribution Amount")
    worksheet.cell(row=total_row_num, column=4, value=f"{total_amount}")

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=contribution_details_{group_id}_{date_format(now(), "YmdHis")}.xlsx'
    workbook.save(response)

    return response






def get_member_loans_and_payments(member_id, group_id):
    # Get the group and member
    group = get_object_or_404(Group, id=group_id)
    member = get_object_or_404(Members, id=member_id, groups=group)

    # Get all loans in the group for the specific member
    member_loans = Loan.objects.filter(member__id=member_id, groups__id=group_id)

    # Get all payments for the member's loans
    member_payments = Payment.objects.filter(loan__member__id=member_id, loan__groups__id=group_id)

    # Calculate the total amount borrowed by the member
    total_amount_to_return = sum([loan.calculate_total_amount() for loan in member_loans])

    # Calculate the total amount paid by the member for their loans
    total_amount_paid = sum([payment.amount_paid for payment in member_payments])

    # Filter loans that are partially paid (loan_status=3)
    partially_paid_loans = member_loans.filter(loan_status=3)

    # Return the loans, total amount to return, payments, and total amount paid
    return partially_paid_loans, total_amount_to_return, member_payments, total_amount_paid


@login_required(login_url='account_login')
def view_member_partially_paid_loans(request, group_id, member_id):
    groups = get_object_or_404(Group, id=group_id)

    # Get specific member's partially paid loans, total amount to return, payments, and total amount paid
    partially_paid_loans, total_amount_to_return, member_payments, total_amount_paid = get_member_loans_and_payments(member_id=member_id, group_id=group_id)

    # Other logic and context setup for your view

    context = {
        'groups': groups,
        'partially_paid_loans': partially_paid_loans,
        'total_amount_to_return': total_amount_to_return,
        'member_payments': member_payments,
        'total_amount_paid': total_amount_paid,
        # Include other relevant context variables
    }

    return render(request, "Main/Payment_loan.html", context)









def download_payment(request, pk):
    groups = get_object_or_404(Group, id=pk)
    member = Members.objects.filter(user=request.user, groups=groups).first()
    member_loans, _ = get_member_loans_and_total_amount(member_id=member.id, group_id=pk)

    # Create a DataFrame to store payment details
    data = {'User First Name': [], 'User Last Name': [], 'Amount Paid (Ksh)': [], 'Date Paid': []}

    for loan in member_loans:
        payments = Payment.objects.filter(loan=loan)
        for payment in payments:
            data['User First Name'].append(loan.member.user.first_name)
            data['User Last Name'].append(loan.member.user.last_name)
            # Use "Ksh" in the "Amount Paid" column
            data['Amount Paid (Ksh)'].append(f'{payment.amount_paid} Ksh')
            # Convert datetime to string without timezone info
            data['Date Paid'].append(payment.payment_date.strftime('%Y-%m-%d %H:%M:%S'))

    payments_df = pd.DataFrame(data)

    # Create a response object with an Excel attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=loan_payments_{request.user.username}.xlsx'

    # Write the DataFrame to the Excel file
    payments_df.to_excel(response, index=False)

    return response